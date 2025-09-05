import pandas as pd
from openpyxl import load_workbook

year=2025
write_to_shared=True
print("Year:",year)

# Define start and end date for time period
PeriodStartDate = pd.Timestamp(year=year, month=1, day=1)
PeriodEndDate = pd.Timestamp(year=year, month=8, day=29)

# Read in hours data from clockify
if year==2019:
    df = pd.read_csv(r"Files\Clockify Exports\2021-02-03 - Clockify_Detailed_Report_01_01_2019-12_31_2019.csv")
elif year==2020:
    df=pd.read_csv(r"Files\Clockify Exports\2021-06-10 - Clockify_Detailed_Report_01_01_2020-12_31_2020.csv")
elif year==2021:
    df=pd.read_csv(r"Files\Clockify Exports\2022-01-24 - Clockify_Time_Report_Detailed_01_01_2021-12_31_2021.csv")
elif year==2022:
    df=pd.read_csv(r"Files\Clockify Exports\2022-12-23 - Clockify_Time_Report_Detailed_01_01_2022-12_31_2022.csv")
elif year==2023:
    df=pd.read_csv(r"Files\Clockify Exports\2024-01-02 - Clockify_Time_Report_Detailed_01_01_2023-12_31_2023.csv")
elif year == 2024:
    df = pd.read_csv(r"Files\Clockify Exports\2025-06-12 - Clockify_Time_Report_Detailed_01_01_2024-12_31_2024.csv")
elif year == 2025:
    df = pd.read_csv(r"Files\Clockify Exports\2025-09-04 - Clockify_Time_Report_Detailed_01_01_2025-12_31_2025.csv")

df['End Date']=df['End Date'].astype('datetime64[ns]')
df['Start Date']=df['Start Date'].astype('datetime64[ns]')

# Categorize Work type
for i, row in df.iterrows():
    project=row['Project']
    if (project=="Sick") | (project=="Stat Holiday")| (project=="Vacation") | (project=="Office Closed"):
        worktype=project
    else:
        worktype="Work"

    df.at[i,'Work Type']=worktype

# Read in rollover data
rollover=pd.read_csv("Files\\"+str(year-1)+"Rollover.csv")

# Read in vacation allotment data
allottment=pd.read_csv(r"Files\\"+str(year)+"VacationAllotment.csv")

# Read in userpaths
paths=pd.read_csv(r"Files\\userpaths.csv")
paths['Start Date']=pd.to_datetime(paths['Start Date'])
paths=paths.set_index("User").to_dict()

# Obtain unique users
users=df.User.unique()
#print(users)

#['Jskow','Afraser','Ashwin Abraham','Bayton','Ben Nowell','Dbandstra','Findlay McCormick','Glenna Case','Jaydon Vanselow','Juan Rojas','Pedro Petraglia','Ryan Stewart','Samuel Allouche','Shawn Smith','Sherry Sun','Sydney Veldhuis','Tdessein']

# Loop through each user, extract hours summary and save into employee file
for user in [#'Afraser',
                #'Ahmed Abdelmoety',
                #'Ayman Abbas',
                #'Bayton',
                #'Ben Nowell',
                #'Dbandstra',
                #'Findlay McCormick',
                # 'Finn Skow',
                #'Glenna Case',
                'Jaydon Vanselow',
                #'Jskow',
                #'Olamide Gabriel',
                # 'Pedro Petraglia',
                # 'Ryan Stewart',
                # 'Samuel Allouche',
                'Shawn Smith',
                # 'Sherry Sun',
                # 'Sydney Veldhuis',
                # 'Tdessein',
             ]:
    print("\nUser: "+user)

    # Read template file and dates in dates for this year into an empty dataframe
    wb = load_workbook(filename="Templates\\"+str(year)+' Template.xlsx')
    ws = wb[str(year)]

    dfxl = pd.DataFrame({'Dates': pd.Series([], dtype='datetime64[ns]')})
    for row in range(7, 400):
        if pd.notna(ws['A' + str(row)].value):
            dfxl = pd.concat([dfxl, pd.Series({"Dates": pd.to_datetime(ws['A' + str(row)].value)}).to_frame().T],
                             join='outer', ignore_index=True)
            #dfxl = dfxl.append({"Dates": pd.to_datetime(ws['A' + str(row)].value)}, ignore_index=True)

    # Time late of period and user's start date
    UserStartDate=paths['Start Date'][user]
    StartDate=max(PeriodStartDate,UserStartDate)
    EndDate=PeriodEndDate

    if UserStartDate>EndDate:
        print(user," was not yet hired in ",year)
    else:
        # Fill in rollover and allottment
        ws.cell(column=15, row=2, value=rollover[rollover.User == user].Vacation.iloc[0])
        ws.cell(column=14, row=2, value=rollover[rollover.User == user].Banked.iloc[0])
        ws.cell(column=15, row=3, value=allottment[allottment.User == user].Vacation.iloc[0])

        # Filter hours to just this user and time frame
        df_grouped = df[df.User == user]

        # Delete hours requirements from before hiring
        printFlag=0
        for row in range(7, 400):
            if pd.notna(ws['A' + str(row)].value):
                if pd.to_datetime(ws['A'+str(row)].value)<UserStartDate:
                    if printFlag==0:
                        print("Removing required hours before:",UserStartDate,'for',user)
                        printFlag=1
                    ws.cell(column=2,row=row,value=0)
                    ws.cell(column=3, row=row, value=0)
                    ws.cell(column=4, row=row, value=0)

        df_grouped = df_grouped.groupby(['User', 'Work Type', 'Start Date'], as_index=False).agg(
            {'Billable': 'first', 'Duration (decimal)': 'sum', 'End Date': 'last'}).reset_index()

        # Filter to time range of interest
        df_grouped = df_grouped[(df_grouped['Start Date'] >= StartDate) & (df_grouped['End Date'] <= EndDate)]

        dfxl2=dfxl.copy()
        for worktype in ['Stat Holiday','Office Closed','Work','Vacation','Sick']:
            dfxl2=dfxl2.merge(df_grouped[df_grouped['Work Type'] == worktype][['Start Date','Duration (decimal)']],left_on='Dates',right_on='Start Date',how='left').fillna(0).drop(columns=['Start Date']).rename(columns={"Duration (decimal)":worktype})

        for i in range(0,len(dfxl)):
            ws.cell(column=6, row=i+7,value=dfxl2.loc[i,'Stat Holiday'])
            ws.cell(column=7, row=i + 7, value=dfxl2.loc[i, 'Office Closed'])
            ws.cell(column=8, row=i + 7, value=dfxl2.loc[i, 'Work'])
            ws.cell(column=9, row=i + 7, value=dfxl2.loc[i, 'Vacation'])
            ws.cell(column=10, row=i + 7, value=dfxl2.loc[i, 'Sick'])

        # Write rollover,  allotted and user labels
        ws.cell(column=13,row=2,value=str(year-1)+" Rollover")
        ws.cell(column=13, row=3, value=str(year) + " Allotted")
        ws.cell(column=6, row=5, value=user + " Hours")

        # Delete hours after period end date
        printFlag=0
        for row in range(7, 400):
            if pd.notna(ws['A' + str(row)].value):
                if pd.to_datetime(ws['A'+str(row)].value)>EndDate:
                    if printFlag==0:
                        print("Removing hours after:",EndDate,'for',user)
                        printFlag=1
                    for c in list(range(1,17)):
                        ws.cell(column=c,row=row,value="")


        wb.save(filename="Outputs\\"+str(year)+" - "+user+".xlsx")

        # Write to Shared folder
        if write_to_shared==True:
            print("Writing summary to Shared folder")
            wb.save(filename=paths['Path'][user]+"\\" + str(year) + " - " + user + ".xlsx")









